import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
import tkinter.simpledialog as simpledialog
import os
from tkinter import Tk, Frame, Label, Button, messagebox
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import csv
from datetime import datetime
import json
from reportlab.lib.pagesizes import letter
from datetime import datetime
import win32print
import win32api
import os

# Cargar el archivo de Excel
archivo_excel = "productos.xlsx"

def buscar_producto_y_agregar(event=None):
    codigo_barras = entrada_codigo.get()
    if not codigo_barras:
        return

    # Load the Excel file and search for the product
    try:
        wb = load_workbook(archivo_excel)
        hoja = wb.active

        for fila in hoja.iter_rows(min_row=2, values_only=True):  # Skip header
            if str(fila[0]) == codigo_barras:  # Match barcode
                nombre = fila[1]  # Product name
                precio_unitario = float(fila[2])  # Product price

                # Call the function to add the product to the cart
                agregar_producto_al_carrito(codigo_barras, nombre, precio_unitario)
                return

        # If product not found
        messagebox.showerror("Error", "Producto no encontrado")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo leer el archivo Excel: {e}")

def asignar_colores_carrito():
    # Recorre todos los elementos en el carrito y asigna colores alternos
    for index, item in enumerate(carrito.get_children()):
        tag = 'color1' if index % 2 == 0 else 'color2'  # Alternar color
        carrito.item(item, tags=(tag,))

def quitar_del_carrito():
    # Obtener el producto seleccionado
    seleccionado = carrito.selection()
    if not seleccionado:
        messagebox.showwarning("Advertencia", "Por favor, selecciona un producto para quitarlo.")
        return
    
    for item in seleccionado:
        carrito.delete(item)

    # Actualizar el total
    calcular_total()
    
    # Reasignar colores a las filas restantes
    asignar_colores_carrito()


def modificar_cantidad():
    # Obtener el producto seleccionado
    seleccionado = carrito.selection()
    if not seleccionado:
        messagebox.showwarning("Advertencia", "Por favor, selecciona un producto para modificar la cantidad.")
        return

    try:
        nueva_cantidad = int(selector_cantidad.get())  # Obtener la cantidad del Spinbox
        if nueva_cantidad <= 0:
            messagebox.showerror("Error", "La cantidad debe ser mayor a 0.")
            return

        for item in seleccionado:
            valores = carrito.item(item, "values")
            codigo_barras, nombre, precio_unitario, _, _ = valores
            precio_unitario = float(precio_unitario.replace(",", ""))  # Convertir el precio a número
            nuevo_precio_final = nueva_cantidad * precio_unitario

            # Actualizar el producto en el carrito
            carrito.item(item, values=(
                codigo_barras,
                nombre,
                f"{precio_unitario:,.2f}",  # Formatear con 2 decimales
                nueva_cantidad,
                f"{nuevo_precio_final:,.2f}"  # Formatear con 2 decimales
            ))

        # Actualizar el total
        calcular_total()

    except ValueError:
        messagebox.showerror("Error", "Por favor, ingresa una cantidad válida.")


def modificar_precio():
    # Obtener el producto seleccionado
    seleccionado = carrito.selection()
    if not seleccionado:
        messagebox.showwarning("Advertencia", "Por favor, selecciona un producto para modificar el precio.")
        return

    def confirmar_precio():
        entrada_precio = entrada_nuevo_precio.get().strip()  # Eliminar espacios en blanco
        entrada_precio = entrada_precio.replace(",", "")  # Reemplazar cualquier coma por un punto
        try:
            nuevo_precio_final = float(entrada_precio)
            if nuevo_precio_final < 0:
                messagebox.showerror("Error", "El precio debe ser mayor o igual a 0.")
                return

            for item in seleccionado:
                valores = carrito.item(item, "values")
                codigo_barras, nombre, precio_unitario, cantidad, _ = valores
                cantidad = int(cantidad)
                nuevo_precio_unitario = nuevo_precio_final / cantidad if cantidad > 0 else 0

                # Formatear el precio unitario y final con 2 decimales
                carrito.item(item, values=(
                    codigo_barras,
                    nombre,
                    f"{nuevo_precio_unitario:,.2f}",  # Formato con 2 decimales
                    cantidad,
                    f"{nuevo_precio_final:,.2f}"  # Formato con 2 decimales
                ))

            # Actualizar el total
            calcular_total()

            ventana_precio.destroy()
        except ValueError:
            messagebox.showerror("Error", "Por favor, ingresa un monto válido.")

    # Crear una ventana emergente
    ventana_precio = tk.Toplevel(ventana)
    ventana_precio.title("Modificar Precio")

    # Ajustar tamaño y centrar la ventana
    ancho = 300
    alto = 150
    x = ventana.winfo_x() + (ventana.winfo_width() - ancho) // 2
    y = ventana.winfo_y() + (ventana.winfo_height() - alto) // 2
    ventana_precio.geometry(f"{ancho}x{alto}+{x}+{y}")

    tk.Label(ventana_precio, text="Nuevo Precio Final:").pack(pady=20)
    entrada_nuevo_precio = tk.Entry(ventana_precio)
    entrada_nuevo_precio.pack(pady=10)

    boton_confirmar = tk.Button(ventana_precio, text="Aceptar", command=confirmar_precio)
    boton_confirmar.pack(pady=10)


def calcular_total():
    total = 0
    for item in carrito.get_children():
        valores = carrito.item(item, "values")
        total += float(valores[4].replace(",", ""))  # Quitar comas antes de convertir a número
    etiqueta_total.config(text=f"Total: ${total:,.2f}")

def nueva_compra():
    # Limpiar el carrito y reiniciar el total
    carrito.delete(*carrito.get_children())  # Eliminar todos los elementos del carrito
    calcular_total()  # Reiniciar el total a 0

# Crear ventana principal
ventana = tk.Tk()
ventana.title("Punto de Venta")
ventana.state('zoomed')
ventana.config(bg="#e6ffe7")

frame_busqueda = tk.Frame(ventana, bg="#e6ffe7")
frame_busqueda.place(x=210, y=20)

tk.Label(frame_busqueda, bg="#e6ffe7", text="Cantidad:").pack(side=tk.LEFT)
selector_cantidad = tk.Spinbox(frame_busqueda, from_=1, to=9, width=5)
selector_cantidad.pack(side=tk.LEFT, padx=5)

tk.Label(frame_busqueda, bg="#e6ffe7", text="Código de barras:").pack(side=tk.LEFT)
entrada_codigo = tk.Entry(frame_busqueda, width=30)
entrada_codigo.pack(side=tk.LEFT, padx=5)

# Asocia la tecla "Enter" al evento de agregar el producto al carrito
entrada_codigo.bind("<Return>", buscar_producto_y_agregar)  # Aquí se añade la asociación con la tecla "Enter"

boton_buscar = tk.Button(frame_busqueda, text="Agregar al carrito", command=buscar_producto_y_agregar, bg="#018a06", fg="#ffffff")
boton_buscar.pack(side=tk.LEFT)


# Tabla del carrito
frame_carrito = tk.Frame(ventana, bg="#e6ffe7")
frame_carrito.place(x=20, y=70)

columnas = ("codigo", "nombre", "precio_unitario", "cantidad", "precio_final")
carrito = ttk.Treeview(frame_carrito, columns=columnas, show="headings", height=22)

# Configuración de colores alternos
carrito.tag_configure('color1', background='#d4edda')  # Color verde claro
carrito.tag_configure('color2', background='#ffffff')   # Color blanco

carrito.heading("codigo", text="Código")
carrito.heading("nombre", text="Nombre")
carrito.heading("precio_unitario", text="Precio Unitario")
carrito.heading("cantidad", text="Cantidad")
carrito.heading("precio_final", text="Precio Final")

# Centrar las columnas
for col in columnas:
    carrito.column(col, anchor="center")

carrito.pack()

contador = 0

# Botones para acciones
frame_botones_categoria = tk.Frame(ventana, bg="#e6ffe7")
frame_botones_categoria.place(x=120, y=551)

def buscar_por_descripcion():
    # Crear ventana emergente para la búsqueda
    ventana_busqueda = tk.Toplevel(ventana)
    ventana_busqueda.title("Buscar Producto")

    # Ajustar tamaño y centrar la ventana
    ancho = 400
    alto = 450
    x = ventana.winfo_x() + (ventana.winfo_width() - ancho) // 2
    y = ventana.winfo_y() + (ventana.winfo_height() - alto) // 2
    ventana_busqueda.geometry(f"{ancho}x{alto}+{x}+{y}")

    # Campo de búsqueda
    tk.Label(ventana_busqueda, text="Buscar por Código o Descripción:").pack(pady=10)
    entrada_busqueda = tk.Entry(ventana_busqueda, width=30)
    entrada_busqueda.pack(pady=10)

    # Crear Treeview para mostrar los resultados
    columnas_busqueda = ("codigo", "nombre", "precio_unitario")
    treeview_busqueda = ttk.Treeview(ventana_busqueda, columns=columnas_busqueda, show="headings", height=10)
    treeview_busqueda.heading("codigo", text="Código")
    treeview_busqueda.heading("nombre", text="Nombre")
    treeview_busqueda.heading("precio_unitario", text="Precio Unitario")

    # Centrar las columnas
    for col in columnas_busqueda:
        treeview_busqueda.column(col, anchor="center")

    treeview_busqueda.pack(pady=10)

    # Definir estilos para los colores
    treeview_busqueda.tag_configure('color1', background='#f0f0f0')  # Color claro
    treeview_busqueda.tag_configure('color2', background='#ffffff')  # Color blanco

    # Contador para alternar colores
    contador = 0

    def actualizar_resultados(event=None):
        query = entrada_busqueda.get().strip().lower()
        if not query:
            for item in treeview_busqueda.get_children():
                treeview_busqueda.delete(item)
            return
        
        try:
            for item in treeview_busqueda.get_children():
                treeview_busqueda.delete(item)

            wb = load_workbook(archivo_excel)
            hoja = wb.active

            for fila in hoja.iter_rows(min_row=2, values_only=True):  # Saltar encabezado
                if len(fila) >= 3:
                    codigo_barras, nombre, precio_unitario = fila[:3]
                    if query in str(codigo_barras).lower() or query in str(nombre).lower():
                        precio_unitario_formateado = f"{float(precio_unitario):,.2f}" if precio_unitario else "0.00"
                        treeview_busqueda.insert("", "end", values=(codigo_barras, nombre, precio_unitario_formateado))
                    
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo realizar la búsqueda: {e}")

    entrada_busqueda.bind("<KeyRelease>", actualizar_resultados)

    # Función para agregar producto al carrito
    def agregar_producto_al_carrito(codigo_barras, nombre, precio_unitario):
        global contador  # Use a global counter for alternating row colors

    # Asegúrate de que el precio unitario sea un float
    precio_unitario = float(precio_unitario)

    # Aquí no verificamos si el producto ya existe en el carrito
    # Simplemente lo agregamos como un nuevo producto
    tag = 'color1' if contador % 2 == 0 else 'color2'  # Alternar color
    carrito.insert("", "end", values=(
        codigo_barras,
        nombre,
        f"{precio_unitario:,.2f}",
        1,  # Comenzar con una cantidad de 1
        f"{precio_unitario:,.2f}"
    ), tags=(tag,))

    # Incrementar el contador para el color de la fila
    contador += 1

    # Actualizar el total
    calcular_total()

    # Botón para agregar el producto seleccionado
    boton_agregar = tk.Button(ventana_busqueda, text="Agregar al Carrito", command=agregar_producto_al_carrito)
    boton_agregar.pack(pady=10)

    # Botón para cerrar la ventana de búsqueda
    boton_cerrar = tk.Button(ventana_busqueda, text="Cerrar", command=ventana_busqueda.destroy)
    boton_cerrar.pack(pady=10)

    # Mostrar la ventana de búsqueda
    ventana_busqueda.mainloop()

# Crear botón en la ventana principal para abrir la ventana de búsqueda
boton_buscar_desc = tk.Button(ventana, text="Buscar", command=buscar_por_descripcion, bg="#018a06", fg="#ffffff")
boton_buscar_desc.place(x=725, y=20)

frame_botones_acciones = tk.Frame(ventana)
frame_botones_acciones.place(x=100, y=100)

frame_total = tk.Frame(ventana, bg="#e6ffe7")
frame_total.place(x=1075, y=50)

boton_quitar = tk.Button(ventana, text="Quitar del carrito", height=2, width=20, command=quitar_del_carrito, bg="#ff4d4d", fg="#ffffff")
boton_quitar.place(x=1115, y=300)

boton_modificar_cantidad = tk.Button(ventana, text="Modificar Cantidad", height=2, width=20, command=modificar_cantidad, bg="#018a06", fg="#ffffff")
boton_modificar_cantidad.place(x=1115, y=360)

boton_modificar_precio = tk.Button(ventana, text="Modificar Precio",height=2, width=20, command=modificar_precio, bg="#018a06", fg="#ffffff")
boton_modificar_precio.place(x=1115, y=420)

boton_nueva_compra = tk.Button(ventana, text="Nueva Compra", command=nueva_compra, bg="#018a06", fg="#ffffff")
boton_nueva_compra.place(x=1200, y=150)

# Etiqueta para mostrar el total
etiqueta_total = tk.Label(frame_total, bg="#e6ffe7", text="Total: $0.00", font=("Arial", 14))
etiqueta_total.pack(pady=10)

def agregar_producto_sin_codigo(nombre_categoria):
    global contador
    # Crear ventana emergente para ingresar el precio
    ventana_precio = tk.Toplevel(ventana)
    ventana_precio.title(f"Ingresar Precio de {nombre_categoria}")

    # Ajustar tamaño y centrar la ventana
    ancho = 300
    alto = 150
    x = ventana.winfo_x() + (ventana.winfo_width() - ancho) // 2
    y = ventana.winfo_y() + (ventana.winfo_height() - alto) // 2
    ventana_precio.geometry(f"{ancho}x{alto}+{x}+{y}")

    tk.Label(ventana_precio , text="Precio del Producto:").pack(pady=20)
    entrada_precio = tk.Entry(ventana_precio)
    entrada_precio.pack(pady=10)

    def agregar_al_carrito():
        global contador
        # Obtener el precio ingresado
        precio = entrada_precio.get().strip()
        try:
            precio = float(precio.replace(",", ""))  # Convertir a número
            if precio < 0:
                messagebox.showerror("Error", "El precio debe ser mayor o igual a 0.")
                return

            # Agregar el producto al carrito con cantidad 1
            tag = 'color1' if contador % 2 == 0 else 'color2'  # Alternar color
            carrito.insert("", "end", values=(nombre_categoria, nombre_categoria, f"{precio:,.2f}", 1, f"{precio:,.2f}"), tags=(tag,))

            # Incrementar el contador
            contador += 1

            # Actualizar el total
            calcular_total()

            # Cerrar la ventana emergente
            ventana_precio.destroy()

        except ValueError:
            messagebox.showerror("Error", "Por favor, ingresa un monto válido.")

    # Botón para confirmar el precio
    boton_agregar = tk.Button(ventana_precio, text="Agregar al Carrito", command=agregar_al_carrito)
    boton_agregar.pack(pady=10)

# El resto de tu código permanece igual


# Crear las filas de botones
botones_textos = [
    "Panificación", "Fiambres y Lácteos", "Limpieza", "Carnes y Verduras", "Útiles y Regalería",
    "Bebidas", "Almacén", "Mascotas", "Promos", "Varios"
]

# Crear botones en dos filas (5 botones en cada fila)
for i, texto in enumerate(botones_textos):
    boton = tk.Button(frame_botones_categoria, text=texto, bg="#018a06", fg="#ffffff", width=20, height=2, command=lambda texto=texto: agregar_producto_sin_codigo(texto))
    boton.grid(row=i//5, column=i%5, padx=5, pady=5)


def consultar_precio():
    # Crear ventana emergente para la búsqueda de precio
    ventana_busqueda_precio = tk.Toplevel(ventana)
    ventana_busqueda_precio.title("Consultar Precio del Producto")

    # Ajustar tamaño y centrar la ventana
    ancho = 600
    alto = 400
    x = ventana.winfo_x() + (ventana.winfo_width() - ancho) // 2
    y = ventana.winfo_y() + (ventana.winfo_height() - alto) // 2
    ventana_busqueda_precio.geometry(f"{ancho}x{alto}+{x}+{y}")

    # Campo de búsqueda
    tk.Label(ventana_busqueda_precio, text="Buscar por Código o Descripción:").pack(pady=10)
    entrada_busqueda_precio = tk.Entry(ventana_busqueda_precio, width=50)
    entrada_busqueda_precio.pack(pady=10)

    # Crear Treeview para mostrar las sugerencias
    treeview = ttk.Treeview(ventana_busqueda_precio, columns=("Código", "Descripción", "Precio", "Stock Min", "Stock Actual", "Sección"), show="headings")
    treeview.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # Definir las columnas
    treeview.heading("Código", text="Código")
    treeview.heading("Descripción", text="Descripción")
    treeview.heading("Precio", text="Precio")
    treeview.heading("Stock Min", text="Stock Min")
    treeview.heading("Stock Actual", text="Stock Actual")
    treeview.heading("Sección", text="Sección")

    # Ajustar el ancho de las columnas
    treeview.column("Código", width=100, anchor="center")
    treeview.column("Descripción", width=200, anchor="center")
    treeview.column("Precio", width=100, anchor="center")
    treeview.column("Stock Min", width=100, anchor="center")
    treeview.column("Stock Actual", width=100, anchor="center")
    treeview.column("Sección", width=100, anchor="center")

    # Función para actualizar las sugerencias en el Treeview
    def actualizar_sugerencias(event=None):
        query = entrada_busqueda_precio.get().strip().lower()
        if not query:
            # Limpiar el Treeview si no hay texto en el campo de búsqueda
            for item in treeview.get_children():
                treeview.delete(item)
            return

        try:
            # Limpiar el Treeview antes de agregar nuevos resultados
            for item in treeview.get_children():
                treeview.delete(item)

            # Buscar en el archivo Excel
            wb = load_workbook(archivo_excel)
            hoja = wb.active

            for fila in hoja.iter_rows(min_row=2, values_only=True):  # Saltar encabezado
                if len(fila) >= 6:
                    codigo_barras, descripcion, precio_unitario, stock_min, stock_actual, seccion = fila[:6]  # Extraer las columnas relevantes

                    # Verificar si el texto ingresado está en el código o la descripción
                    if query in str(codigo_barras).lower() or query in str(descripcion).lower():
                        # Agregar la fila al Treeview
                        precio_formateado = f"${float(precio_unitario):,.2f}"
                        # Aquí los valores se insertan en el orden correcto
                        treeview.insert("", "end", values=(codigo_barras, descripcion, precio_formateado, stock_min, stock_actual, seccion))

        except Exception as e:
            messagebox.showerror("Error", f"Hubo un problema al buscar los productos: {e}")

    # Asociar la actualización de sugerencias al evento de escribir en el campo de búsqueda
    entrada_busqueda_precio.bind("<KeyRelease>", actualizar_sugerencias)

    # Botón para cerrar la ventana de búsqueda de precios
    boton_cerrar = tk.Button(ventana_busqueda_precio, text="Cerrar", command=ventana_busqueda_precio.destroy)
    boton_cerrar.pack(pady=10)


    # Función para mostrar el producto seleccionado al hacer clic en una fila del Treeview
    def seleccionar_producto(event):
        item = treeview.selection()  # Obtener la fila seleccionada
        if item:
            values = treeview.item(item)["values"]
            codigo_barras, descripcion, precio_unitario, stock_min, stock_actual, seccion = values

            # Mostrar la información del producto seleccionado
            mensaje = f"Producto: {descripcion}\nCódigo: {codigo_barras}\nStock Actual: {stock_actual}\nStock Min: {stock_min}\nPrecio: {precio_unitario}\nSección: {seccion}"
            messagebox.showinfo("Producto Seleccionado", mensaje)

    # Asociar el clic en una fila del Treeview a la función de seleccionar el producto
    treeview.bind("<Double-1>", seleccionar_producto)

# Crear botón en la ventana principal para abrir la ventana de consulta de precio
boton_consultar_precio = tk.Button(ventana, text="Consultar Precio", command=consultar_precio, bg="#018a06", fg="#ffffff")
boton_consultar_precio.place(x=1090, y=150)




def cargar_configuracion():
    # Cargar la configuración desde el archivo JSON
    config_path = os.path.join("user_data", "configuracion.json")
    with open(config_path, 'r') as f:
        return json.load(f)

def registrar_venta(productos_vendidos, total, monto_efectivo, monto_transferencia):
    # Crear o abrir el archivo de ventas
    with open('ventas.csv', mode='a', newline='') as file:
        writer = csv.writer(file)
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # Escribir la venta en el archivo, incluyendo efectivo y transferencia
        writer.writerow([fecha, total, str(productos_vendidos), monto_efectivo, monto_transferencia])



def imprimir_ticket(ticket_path):
    """Función para imprimir el ticket generado automáticamente."""
    try:
        if os.path.exists(ticket_path):
            # Imprimir el ticket sin mostrar el diálogo
            win32api.ShellExecute(0, "print", ticket_path, None, ".", 0)
            messagebox.showinfo("Éxito", "El ticket se está imprimiendo.")
        else:
            messagebox.showerror("Error", "El ticket no se generó correctamente.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo imprimir el ticket: {e}")

def generar_ticket_pdf(productos_vendidos, total):
    try:
        # Cargar la configuración
        config = cargar_configuracion()
        nombre_negocio = config["nombre_negocio"]
        logo_path = config["logo"]
        mensaje_final = config["mensaje_final"]

        # Crear la carpeta 'ventas' si no existe
        carpeta_ventas = os.path.join(os.path.expanduser("~"), "Documents", "Ventas")
        os.makedirs(carpeta_ventas, exist_ok=True)

        # Generar el nombre del archivo basado en la fecha y hora
        fecha_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(carpeta_ventas, f"ticket_{fecha_hora}.pdf")

        width_mm = 58
        width = width_mm * 2.83464567
        
        # Calcular la altura del ticket
        altura_base = 90  # altura mínima base
        altura_por_linea = 10  # altura por cada línea de producto
        lineas = len(productos_vendidos) + 8  # 6 líneas para encabezados, total y mensaje final
        height = altura_base + (lineas * altura_por_linea)

        margen_izquierdo_mm = 0
        margen_derecho_mm = 0
        margen_izquierdo = margen_izquierdo_mm * 1
        margen_derecho = margen_derecho_mm * 1

        c = canvas.Canvas(file_path, pagesize=(width, height))

        # Agregar el logo centrado
        if os.path.exists(logo_path):
            image_width = width * 0.3
            image_height = image_width
            x_center = (width - image_width) / 2
            y_position_image = height - image_height - 10
            c.drawImage(ImageReader(logo_path), x_center, y_position_image, 
                        width=image_width, height=image_height, preserveAspectRatio=True, anchor='nw')
        else:
            raise FileNotFoundError("No se encontró la imagen del logo.")

        # Agregar el nombre del negocio
        c.setFont("Helvetica-Bold", 10)
        c.drawString(margen_izquierdo, y_position_image - 20, nombre_negocio)

        # Agregar la fecha y hora
        c.setFont("Helvetica", 8)
        fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c.drawString(margen_izquierdo, y_position_image - 35, f"Fecha: {fecha_hora}")

        # Encabezado de columnas
        y_position = y_position_image - 50
        c.setFont("Helvetica-Bold", 8)
        c.drawString(margen_izquierdo, y_position, "Producto")
        c.drawString(margen_izquierdo + 84, y_position, "Cant.")  # Ajusta la posición para centrar
        c.drawString(width - margen_derecho - 40, y_position, "Importe")

        # Línea de separación
        y_position -= 10
        c.line(margen_izquierdo, y_position, width - margen_derecho, y_position)

        # Agregar productos
        y_position -= 10
        c.setFont("Helvetica", 8)  # Cambiar a una fuente más legible
        for codigo_barras, cantidad, precio_unitario in productos_vendidos:
            for item in carrito.get_children():
                valores = carrito.item(item, "values")
                if valores[0] == codigo_barras:
                    nombre = valores[1]
                    if len(nombre) > 13:
                        nombre = nombre[:17] + "..."  # Truncar el nombre si es muy largo
                    importe = cantidad * precio_unitario
                    c.drawString(margen_izquierdo, y_position, nombre)
                    c.drawString(margen_izquierdo + 90, y_position, str(cantidad))
                    c.drawString(width - margen_derecho - 40, y_position, f"${importe:,.2f}")
                    y_position -= 10  # Mover hacia abajo para la siguiente línea

        # Línea de separación antes del total
        c.line(margen_izquierdo, y_position, width - margen_derecho, y_position)
        y_position -= 10
        c.setFont("Helvetica-Bold", 9)
        c.drawString(margen_izquierdo, y_position, "TOTAL:")
        c.drawString(width - margen_derecho - 50, y_position, f"${total:,.2f}")

        # Mensaje final
        y_position -= 20
        c.setFont("Helvetica", 8)
        c.drawString(margen_izquierdo, y_position, mensaje_final)

        # Finalizar el PDF
        c.showPage()
        c.save()

        return file_path  # Retornar la ruta del archivo generado

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo generar el ticket: {e}")

def cobrar_carrito():
    total = 0
    productos_vendidos = []  # Lista para almacenar los productos vendidos
    stock_insuficiente = False  # Variable para verificar si hay stock insuficiente

    for item in carrito.get_children():
        valores = carrito.item(item, "values")
        total += float(valores[4].replace(",", ""))  # Obtener el precio final de cada producto
        codigo_barras = valores[0]
        cantidad_vendida = int(valores[3])  # Obtener la cantidad vendida
        precio_unitario = float(valores[2].replace(",", ""))  # Obtener el precio unitario

        # Verificar stock antes de cobrar
        try:
            wb = load_workbook(archivo_excel)
            hoja = wb.active
            for fila in hoja.iter_rows(min_row=2):  # Saltar encabezado
                if str(fila[0].value) == codigo_barras:  # Comparar con el código de barras
                    stock_actual = fila[4].value  # Suponiendo que el stock actual está en la columna 5 (índice 4)

                    # Asegúrate de convertir stock_actual a un número
                    if isinstance(stock_actual, str):
                        stock_actual = int(stock_actual)  # Convertir a entero si es una cadena

                    if cantidad_vendida > stock_actual:
                        stock_insuficiente = True  # Marcar que hay stock insuficiente
                        messagebox.showerror("Error", f"No hay suficiente stock para el producto: {valores[1]}. "
                                                       f"Stock actual: {stock_actual}, Cantidad solicitada: {cantidad_vendida}.")
                        break  # Salir del bucle si hay stock insuficiente

            if stock_insuficiente:
                break  # Salir del bucle principal si se detectó stock insuficiente

            # Verificar si el producto ya está en productos_vendidos
            for producto in productos_vendidos:
                if producto[0] == codigo_barras:
                    # Si ya existe, solo actualizar la cantidad
                    producto[1] += cantidad_vendida
                    break
            else:
                # Si no existe, agregar el producto a la lista de productos vendidos
                productos_vendidos.append([codigo_barras, cantidad_vendida, precio_unitario])  # Agregar el precio unitario

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo verificar el stock: {e}")
            return

    if stock_insuficiente:
        return  # No continuar si hay stock insuficiente

    # Mostrar el total al usuario y pedir los montos de efectivo y transferencia
    ventana_pago = tk.Toplevel(ventana)
    ventana_pago.title("Cobro")

    tk.Label(ventana_pago, text=f"Total a pagar: ${total:,.2f}").pack(pady=10)

    tk.Label(ventana_pago, text="Monto en Efectivo:").pack(pady=5)
    entrada_efectivo = tk.Entry(ventana_pago)
    entrada_efectivo.insert(0, "0")  # Valor inicial de 0
    entrada_efectivo.pack(pady=5)

    tk.Label(ventana_pago, text="Monto por Transferencia:").pack(pady=5)
    entrada_transferencia = tk.Entry(ventana_pago)
    entrada_transferencia.insert(0, "0")  # Valor inicial de 0
    entrada_transferencia.pack(pady=5)

    def confirmar_pago():
        try:
            monto_efectivo = float(entrada_efectivo.get())
            monto_transferencia = float(entrada_transferencia.get())
            monto_total_recibido = monto_efectivo + monto_transferencia

            if monto_total_recibido < total:
                messagebox.showerror("Error", "El monto total recibido es menor al total a pagar. Intente de nuevo.")
                return
            vuelto = monto_total_recibido - total
            messagebox.showinfo("Vuelto", f"El vuelto es: ${vuelto:,.2f}")

            try:
                wb = load_workbook(archivo_excel)
                hoja = wb.active

                for codigo_barras, cantidad_vendida, _ in productos_vendidos:
                    for fila in hoja.iter_rows(min_row=2):  # Saltar encabezado
                        if str(fila[0].value) == codigo_barras:
                            stock_actual = fila[4].value

                            if isinstance(stock_actual, str):
                                stock_actual = int(stock_actual)

                            nuevo_stock = stock_actual - cantidad_vendida
                            fila[4].value = nuevo_stock
                            break
                wb.save(archivo_excel)
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo actualizar el stock: {e}")
            
            registrar_venta(productos_vendidos, total, monto_efectivo, monto_transferencia)
            ticket_path = generar_ticket_pdf(productos_vendidos, total)

            print(f"Ruta del ticket generado: {ticket_path}")  # Línea de depuración

            if messagebox.askyesno("Imprimir Ticket", "¿Desea imprimir el ticket generado?"):
                imprimir_ticket(ticket_path)  # Llama a la nueva función para imprimir

            nueva_compra()
            ventana_pago.destroy()

        except ValueError:
            messagebox.showerror("Error", "Por favor, ingresa montos válidos.")

    boton_confirmar = tk.Button(ventana_pago, text="Confirmar Pago", command=confirmar_pago)
    boton_confirmar.pack(pady=10)

boton_cobrar = tk.Button(frame_total, text="Cobrar", width=30, command=cobrar_carrito, bg="#018a06", fg="#ffffff")
boton_cobrar.pack(side=tk.TOP, padx=5, pady=5)



def consultar_ventas():
    # Crear ventana emergente para consultar ventas
    ventana_consulta = tk.Toplevel(ventana)
    ventana_consulta.title("Consultar Ventas")

    # Ajustar tamaño y centrar la ventana
    ancho = 700  # Aumentar el ancho para más columnas
    alto = 400
    x = ventana.winfo_x() + (ventana.winfo_width() - ancho) // 2
    y = ventana.winfo_y() + (ventana.winfo_height() - alto) // 2
    ventana_consulta.geometry(f"{ancho}x{alto}+{x}+{y}")

    # Crear Treeview para mostrar las ventas
    columnas_ventas = ("Fecha", "Total", "Productos", "Efectivo", "Transferencia")
    treeview_ventas = ttk.Treeview(ventana_consulta, columns=columnas_ventas, show="headings", height=15)
    treeview_ventas.heading("Fecha", text="Fecha")
    treeview_ventas.heading("Total", text="Total")
    treeview_ventas.heading("Productos", text="Productos")
    treeview_ventas.heading("Efectivo", text="Efectivo")
    treeview_ventas.heading("Transferencia", text="Transferencia")

    treeview_ventas.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # Cargar las ventas desde el archivo CSV
    try:
        with open('ventas.csv', mode='r') as file:
            reader = csv.reader(file)
            for row in reader:
                treeview_ventas.insert("", "end", values=row)
    except FileNotFoundError:
        messagebox.showerror("Error", "No se encontró el archivo de ventas.")



from datetime import datetime, timedelta

menu_bar = tk.Menu(ventana)
menu_ventas = tk.Menu(menu_bar, tearoff=0)
menu_ventas.add_command(label="Consultar Ventas", command=consultar_ventas)

menu_bar.add_cascade(label="Ventas", menu=menu_ventas)

def calcular_recaudacion(periodo):
    total_recaudacion = 0
    total_efectivo = 0
    total_transferencia = 0
    cantidad_ventas = 0  # Inicializar contador de ventas

    try:
        with open('ventas.csv', mode='r') as file:
            reader = csv.reader(file)
            for row in reader:
                # Asegúrate de que la fila tiene al menos 5 columnas
                if len(row) < 5:
                    continue  # Saltar filas que no tienen suficientes columnas

                fecha_str = row[0]
                total = float(row[1])
                efectivo = float(row[3])  # Efectivo
                transferencia = float(row[4])  # Transferencia
                fecha = datetime.strptime(fecha_str, "%Y-%m-%d %H:%M:%S")

                # Contar la venta si está dentro del periodo
                if (periodo == 'dia' and fecha.date() == datetime.now().date()) or \
                   (periodo == 'semana' and fecha >= datetime.now() - timedelta(days=7)) or \
                   (periodo == 'mes' and fecha.month == datetime.now().month and fecha.year == datetime.now().year) or \
                   (periodo == 'anio' and fecha.year == datetime.now().year):
                    total_recaudacion += total
                    total_efectivo += efectivo
                    total_transferencia += transferencia
                    cantidad_ventas += 1  # Incrementar contador de ventas

        # Mostrar la información de recaudación y cantidad de ventas
        messagebox.showinfo("Recaudación", f"La recaudación total para el periodo seleccionado es: ${total_recaudacion:,.2f}\n"
                                           f"Efectivo: ${total_efectivo:,.2f}\n"
                                           f"Transferencia: ${total_transferencia:,.2f}\n"
                                           f"Cantidad de Ventas: {cantidad_ventas}")

    except FileNotFoundError:
        messagebox.showerror("Error", "No se encontró el archivo de ventas.")

def ver_recaudacion():
    # Crear ventana emergente para ver recaudación
    ventana_recaudacion = tk.Toplevel(ventana)
    ventana_recaudacion.title("Recaudación")

    # Ajustar tamaño y centrar la ventana
    ancho = 400
    alto = 300
    x = ventana.winfo_x() + (ventana.winfo_width() - ancho) // 2
    y = ventana.winfo_y() + (ventana.winfo_height() - alto) // 2
    ventana_recaudacion.geometry(f"{ancho}x{alto}+{x}+{y}")

    # Crear botones para seleccionar el periodo
    boton_dia = tk.Button(ventana_recaudacion, text="Recaudación del Día", command=lambda: calcular_recaudacion('dia'))
    boton_dia.pack(pady=10)

    boton_semana = tk.Button(ventana_recaudacion, text="Recaudación de la Semana", command=lambda: calcular_recaudacion('semana'))
    boton_semana.pack(pady=10)

    boton_mes = tk.Button(ventana_recaudacion, text="Recaudación del Mes", command=lambda: calcular_recaudacion('mes'))
    boton_mes.pack(pady=10)

    boton_anio = tk.Button(ventana_recaudacion, text="Recaudación del Año", command=lambda: calcular_recaudacion('anio'))
    boton_anio.pack(pady=10)



menu_ventas.add_command(label="Ver Recaudación", command=ver_recaudacion)
ventana.config(menu=menu_bar)
frame_busqueda = tk.Frame(ventana, bg="#e6ffe7")
frame_busqueda.pack(pady=10)


# Botones para acciones


ventana.mainloop()